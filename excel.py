from openpyxl import Workbook, load_workbook
from word import *
import os
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats, linalg
from scipy.stats import pearsonr, ttest_ind, ttest_1samp

def write_sense_list(word):

    if word.length == 0:
        print("Sense List not possible for given input: " + word.label)
        exit()

    wb = Workbook()
    ws1 = wb.active

    ws1['A1'], ws1['B1'], ws1['C1'], ws1['D1'], ws1['E1'], ws1['F1'], ws1['G1'] = "Raw Date", "Start Date", "End Date", "Word Form", "Identifiers", "Categories", "PoS"
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 110
    ws1.column_dimensions['F'].width = 35

    for num in range(len(word.senses)):
        ws1.cell(row=num+2, column=1).value = word.senses[num].raw_date
        ws1.cell(row=num+2, column=2).value = word.senses[num].date
        if word.senses[num].end_date != 3000:
            ws1.cell(row=num+2, column=3).value = word.senses[num].end_date
        ws1.cell(row=num+2, column=4).value = word.senses[num].word_form
        ws1.cell(row=num+2, column=5).value = word.senses[num].identifiers
        ws1.cell(row=num+2, column=6).value = word.senses[num].categories
        ws1.cell(row=num+2, column=7).value = word.senses[num].PoS
        #print(word.senses[num].listed_cat)

    save_loc = os.getcwd() + "[" + word.PoS + "] " + word.label + ".xlsx"
    wb.save(save_loc)
    return save_loc